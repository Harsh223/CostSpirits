import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from io import BytesIO
import re
import ast
import os
import json
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode
import uuid
import time

# Define predefined subsystems from the reference image
AVAILABLE_SUBSYSTEMS = {
    "Structural/Mechanical Group": [
        "Structure",
        "Mechanisms"
    ],
    "Electrical Power & Distribution Group": [
        "Electrical Power",
        "Power Dist/Reg/Cont"
    ],
    "CC&DH Group": [
        "Data Management",
        "Communication",
        "Antennas",
        "Instrumentation Display & Control"
    ],
    "Other Subsystems": [
        "Avionics",
        "ASE",
        "Range Safety",
        "Separation",
        "Thermal Control",
        "Crew Accommodations",
        "ECLS",
        "Launch & Landing Safety",
        "Miscellaneous",
        "Attitude Control/GN&C",
        "Engines",
        "Propulsion",
        "Reaction Control",
        "Solid/Kick Motor",
        "Thrust Vector Control"
    ]
}

# Load custom headers for each subsystem from dict.txt
def load_subsystem_headers():
    json_path = os.path.join(os.path.dirname(__file__), 'subsystem_headers.json')
    if os.path.exists(json_path):
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            # Convert list of {name, headers} to dict: name -> headers
            return {entry['name']: entry['headers'] for entry in data if 'name' in entry and 'headers' in entry}
    return {}

SUBSYSTEM_HEADERS = load_subsystem_headers()

def sanitize_sheet_name(name):
    # Excel sheet names cannot contain: : \ / ? * [ ]
    return re.sub(r'[:\\/?*\[\]]', ' ', name)[:31]  # Also limit to 31 chars

# Helper to create a styled Excel workbook
def create_template(subsystems):
    wb = Workbook()
    wb.properties.creator = "Harsh Kumar"
    # Remove default sheet
    wb.remove(wb.active)
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    for subsystem in subsystems:
        safe_name = sanitize_sheet_name(subsystem)
        # Try exact match, then fallback to group name match (case-insensitive, ignoring common variations)
        headers = SUBSYSTEM_HEADERS.get(safe_name)
        if headers is None:
            # Try to match group names in a more robust way
            for key in SUBSYSTEM_HEADERS:
                if key.replace('&', 'and').replace('/', '').replace(' ', '').lower() == safe_name.replace('&', 'and').replace('/', '').replace(' ', '').lower():
                    headers = SUBSYSTEM_HEADERS[key]
                    break
        if headers is None:
            headers = []
        ws = wb.create_sheet(title=safe_name)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        ws.freeze_panes = ws["A2"]
    return wb

def create_mass_budget_template(uploaded_file):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.chart import BarChart, Reference
    import pandas as pd
    import random
    xls = pd.ExcelFile(uploaded_file)
    subsystems = xls.sheet_names
    subsystem_wbs = {}
    for sheet in subsystems:
        df = pd.read_excel(xls, sheet_name=sheet)
        wbs_col = None
        for col in df.columns:
            if col.strip().lower() in ["wbs item", "wbs element", "wbs"]:
                wbs_col = col
                break
        if wbs_col:
            unique_wbs = sorted(df[wbs_col].dropna().unique())
        else:
            unique_wbs = []
        subsystem_wbs[sheet] = unique_wbs
    wb = Workbook()
    wb.properties.creator = "CostSpirits"
    ws_budget = wb.active
    ws_budget.title = "Mass Budget Table"
    ws_budget.append(["Subsystem", "Total Mass (user entry)"])
    for i, subsystem in enumerate(subsystems, 2):
        ws_budget.append([subsystem, 0])
    # Style header for budget sheet (gold)
    budget_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    budget_font = Font(bold=True, color="FFFFFF")
    budget_align = Alignment(horizontal="center", vertical="center")
    for col in range(1, 3):
        cell = ws_budget.cell(row=1, column=col)
        cell.fill = budget_fill
        cell.font = budget_font
        cell.alignment = budget_align
    # Add a bar chart for total mass per subsystem
    chart = BarChart()
    chart.title = "Total Mass by Subsystem"
    chart.y_axis.title = "Total Mass"
    chart.x_axis.title = "Subsystem"
    data = Reference(ws_budget, min_col=2, min_row=1, max_row=1+len(subsystems), max_col=2)
    cats = Reference(ws_budget, min_col=1, min_row=2, max_row=1+len(subsystems))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    ws_budget.add_chart(chart, "D2")
    # Each subsystem sheet with unique header color
    header_colors = ["4F81BD", "C0504D", "9BBB59", "8064A2", "F79646", "2C4D75", "1F497D", "E46C0A", "00B050", "7030A0"]
    for idx, subsystem in enumerate(subsystems):
        ws = wb.create_sheet(title=sanitize_sheet_name(subsystem))
        # Top: total mass (linked to budget table, Excel formula)
        ws["A1"] = "Total Subsystem Mass (should match Mass Budget Table)"
        ws["B1"] = f"=('Mass Budget Table'!B{idx+2})"
        # Header row with unique color
        color = header_colors[idx % len(header_colors)]
        header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")
        ws["A2"] = "WBS"
        ws["B2"] = "Include? (Y/N)"
        ws["C2"] = "Weight (if Y)"
        for col in range(1, 4):
            cell = ws.cell(row=2, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        for i, wbs in enumerate(subsystem_wbs[subsystem], 3):
            ws.cell(row=i, column=1, value=wbs)
            ws.cell(row=i, column=2, value="N")
            ws.cell(row=i, column=3, value="")
        ws.freeze_panes = ws["A3"]
    return wb

# Streamlit app
def main():
    st.set_page_config(page_title="CostSpirits: Subsystem Cost Estimator", layout="wide")
    st.title("CostSpirits: Subsystem Cost Estimator")
    page = st.sidebar.radio("Select Page", ["Configure Calculator Page", "Generate Template", "Cost Analysis"])
    if page == "Generate Template":
        st.header("Step 1: Select Subsystems for Template")
        if 'selected_subsystems' not in st.session_state:
            st.session_state.selected_subsystems = set()
        if 'selected_groups' not in st.session_state:
            st.session_state.selected_groups = set()
        selected_subsystems = set()
        selected_groups = set()
        for group, subsystems in AVAILABLE_SUBSYSTEMS.items():
            with st.expander(f"{group}", expanded=True):
                show_group_checkbox = group != "Other Subsystems"
                if show_group_checkbox:
                    group_checked = st.checkbox(f"Select ALL in {group}", key=f"group_{group}")
                    if group_checked:
                        selected_groups.add(group)
                else:
                    group_checked = False
                st.write("Select subsystems from this group:")
                for subsystem in subsystems:
                    if st.checkbox(subsystem, key=subsystem, value=False if group_checked else None, disabled=group_checked):
                        selected_subsystems.add(subsystem)
        st.session_state.selected_subsystems = selected_subsystems
        st.session_state.selected_groups = selected_groups
        # Prepare final list of sheets: group names if selected, else individual subsystems
        sheets = list(selected_groups) + [s for s in selected_subsystems if not any(s in AVAILABLE_SUBSYSTEMS[g] for g in selected_groups)]
        if len(sheets) > 0:
            if st.button("Generate Template"):
                wb = create_template(sheets)
                bio = BytesIO()
                wb.save(bio)
                st.success(f"Template generated with {len(sheets)} selected sheets! Download below:")
                st.download_button(
                    label="Download Excel Template",
                    data=bio.getvalue(),
                    file_name="CostSpirits_Subsystem_Template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.info("Please select at least one subsystem or group to generate the template.")

    elif page == "Configure Calculator Page":
        st.header("Step 2: Configure Calculator")
        # --- Move Export, Save, Load buttons here ---
        if 'show_upload_modal' not in st.session_state:
            st.session_state.show_upload_modal = True
        if st.session_state.show_upload_modal:
            st.info("Do you already have a filled historical cost data sheet?")
            col1, col2 = st.columns(2)
            if col1.button("Yes, I have the filled sheet", key="modal_yes"):
                st.session_state.show_upload_modal = False
            if col2.button("No, I need to generate the template", key="modal_no"):
                st.session_state.show_upload_modal = False
                st.warning("Please go to the 'Generate Template' page, download the template, fill it with your historical data, and then return here.")
        if not st.session_state.show_upload_modal:
            uploaded = st.file_uploader("Upload Excel file", type=["xlsx"])
            if uploaded:
                st.session_state.uploaded_file = uploaded
                st.success("File uploaded! Visualizing template content below:")
                xls = pd.ExcelFile(uploaded)
                # Prepare table data for all sheets
                table_data = []
                for sheet in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet)
                    mission_col = df.columns[0] if not df.empty else None
                    num_missions = df[mission_col].nunique(dropna=True) if mission_col else 0
                    table_data.append({"Subsystem/Component": sheet, "Number of Missions": num_missions, "Rows": len(df)})
                st.subheader("Subsystem Sheets Found:")
                st.table(table_data)
                st.write("---")
                st.subheader("Preview of Each Subsystem Sheet:")
                for sheet in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet)
                    with st.expander(f"{sheet} ({df[df.columns[0]].nunique(dropna=True)} missions)", expanded=False):
                        st.dataframe(df)
                if st.button("Proceed to Cost Analysis"):
                    st.session_state.page = "Cost Analysis"
                    st.query_params["page"] = "Cost Analysis"
                    st.success("Redirected to Cost Analysis. Please select the tab from the sidebar if not automatically redirected.")
                # Mass Budget Template button ONLY here
                if st.button("Download Mass Budget Excel Template"):
                    wb = create_mass_budget_template(uploaded)
                    bio = BytesIO()
                    wb.save(bio)
                    st.success("Mass Budget Template generated! Download below:")
                    st.download_button(
                        label="Download Mass Budget Excel Template",
                        data=bio.getvalue(),
                        file_name="CostSpirits_Mass_Budget_Template.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        # --- Move all action buttons to the bottom, each in its own subsection ---
        st.markdown("---")
        st.subheader("Export to Excel")
        st.caption("Download an Excel file containing all your cost analysis results and breakdowns. Use this after you have completed your analysis.")
        export_clicked = st.button("Export to Excel", key="export_to_excel")
        if export_clicked:
            uploaded = st.session_state.get('uploaded_file')
            if not uploaded:
                st.warning("Please upload a filled template before exporting to Excel.")
            else:
                import io
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                output = io.BytesIO()
                wb = Workbook()
                wb.remove(wb.active)
                subsystem_results = st.session_state.get('subsystem_results', {})
                # Define styles
                header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                header_align = Alignment(horizontal="center", vertical="center")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                alt_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                # Color palette for sheets
                sheet_colors = ["4F81BD", "C0504D", "9BBB59", "8064A2", "F79646", "2C4D75", "1F497D", "E46C0A", "00B050", "7030A0"]
                for idx, (sheet, result_df) in enumerate(subsystem_results.items()):
                    ws = wb.create_sheet(title=sheet[:31])
                    color = sheet_colors[idx % len(sheet_colors)]
                    sheet_header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    # Write headers
                    for c_idx, col in enumerate(result_df.columns, 1):
                        cell = ws.cell(row=1, column=c_idx, value=col)
                        cell.fill = sheet_header_fill
                        cell.font = header_font
                        cell.alignment = header_align
                        cell.border = thin_border
                    # Write data rows with alternating fill
                    for r_idx, row in enumerate(result_df.reset_index(drop=True).itertuples(index=False), 2):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=value)
                            cell.border = thin_border
                            if r_idx % 2 == 0:
                                cell.fill = alt_fill
                    # Auto-fit columns
                    for col in ws.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
                    # --- Add EUR table below main table ---
                    eur_df = st.session_state.get(f"eur_df_{sheet}")
                    if eur_df is not None:
                        start_row = ws.max_row + 2
                        for c_idx, col in enumerate(eur_df.columns, 1):
                            cell = ws.cell(row=start_row, column=c_idx, value=col)
                            cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                            cell.font = header_font
                            cell.alignment = header_align
                            cell.border = thin_border
                        for r_idx, row in enumerate(eur_df.reset_index(drop=True).itertuples(index=False), start_row+1):
                            for c_idx, value in enumerate(row, 1):
                                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                                cell.border = thin_border
                                if r_idx % 2 == 0:
                                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    # Auto-fit columns again
                    for col in ws.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
                # Total Cost Breakdown sheet (distinct style)
                if 'subsystem_results' in st.session_state:
                    ws = wb.create_sheet(title="Total Cost Breakdown")
                    user_mass_df = st.session_state.get('user_mass_df')
                    infl_df = st.session_state.get('infl_df')
                    breakdown_header_fill = PatternFill(start_color="005fa3", end_color="005fa3", fill_type="solid")
                    breakdown_alt_fill = PatternFill(start_color="E3F0FF", end_color="E3F0FF", fill_type="solid")
                    if user_mass_df is not None:
                        for c_idx, col in enumerate(user_mass_df.columns, 1):
                            cell = ws.cell(row=1, column=c_idx, value=col)
                            cell.fill = breakdown_header_fill
                            cell.font = header_font
                            cell.alignment = header_align
                            cell.border = thin_border
                        for r_idx, row in enumerate(user_mass_df.reset_index(drop=True).itertuples(index=False), 2):
                            for c_idx, value in enumerate(row, 1):
                                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                                cell.border = thin_border
                                if r_idx % 2 == 0:
                                    cell.fill = breakdown_alt_fill
                    if infl_df is not None:
                        start_row = ws.max_row + 2
                        for c_idx, col in enumerate(infl_df.columns, 1):
                            cell = ws.cell(row=start_row, column=c_idx, value=col)
                            cell.fill = breakdown_header_fill
                            cell.font = header_font
                            cell.alignment = header_align
                            cell.border = thin_border
                        for r_idx, row in enumerate(infl_df.reset_index(drop=True).itertuples(index=False), start_row+1):
                            for c_idx, value in enumerate(row, 1):
                                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                                cell.border = thin_border
                                if r_idx % 2 == 0:
                                    cell.fill = breakdown_alt_fill
                    # Auto-fit columns
                    for col in ws.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
                wb.save(output)
                st.download_button(
                    label="Download Cost Analysis Excel",
                    data=output.getvalue(),
                    file_name="CostSpirits_Cost_Analysis.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    elif page == "Cost Analysis":
        st.header("Cost Analysis")
        uploaded = st.session_state.get('uploaded_file')
        if not uploaded:
            st.info("Please upload a filled template first on the 'Upload Filled Template' page.")
            return
        # --- Ensure subsystem_results is initialized ---
        if 'subsystem_results' not in st.session_state:
            st.session_state['subsystem_results'] = {}
        xls = pd.ExcelFile(uploaded)
        tab_names = xls.sheet_names
        tabs = st.tabs(tab_names)
        # --- Per-subsystem Tabs ---
        for i, sheet in enumerate(tab_names):
            with tabs[i]:
                st.subheader(f"Subsystem: {sheet}")
                df = pd.read_excel(xls, sheet_name=sheet)
                # Find unique WBS elements
                wbs_col = None
                for col in df.columns:
                    if col.strip().lower() in ["wbs item", "wbs element", "wbs"]:
                        wbs_col = col
                        break
                if wbs_col is None:
                    st.warning("No WBS column found in this sheet.")
                    continue
                unique_wbs = sorted(df[wbs_col].dropna().unique())
                # Dynamic section numbering
                section_num = 1
                # --- WBS Merging UI ---
                if len(unique_wbs) > 1:
                    st.markdown(f"#### {section_num}. (Optional) Merge WBS Items for Analysis")
                    section_num += 1
                    wbs_merge_dict = {}
                    wbs_to_merge = st.multiselect(
                        "Select WBS items to merge (hold Ctrl to select multiple):",
                        options=unique_wbs,
                        key=f"merge_select_{sheet}"
                    )
                    if wbs_to_merge:
                        merge_name = st.text_input("Enter a common name for the merged WBS group:", key=f"merge_name_{sheet}")
                        if merge_name and st.button("Add Merge Group", key=f"add_merge_{sheet}"):
                            if 'wbs_merge_groups' not in st.session_state:
                                st.session_state['wbs_merge_groups'] = {}
                            if sheet not in st.session_state['wbs_merge_groups']:
                                st.session_state['wbs_merge_groups'][sheet] = {}
                            st.session_state['wbs_merge_groups'][sheet][merge_name] = list(wbs_to_merge)
                            st.success(f"Merged group '{merge_name}' created for: {', '.join(wbs_to_merge)}")
                # Prepare WBS list for checklist (merged + unmerged)
                merge_groups = st.session_state.get('wbs_merge_groups', {}).get(sheet, {})
                merged_wbs_flat = [w for group in merge_groups.values() for w in group]
                wbs_for_checklist = list(merge_groups.keys()) + [w for w in unique_wbs if w not in merged_wbs_flat]
                st.markdown(f"#### {section_num}. Select WBS Components Present in Your Subsystem")
                section_num += 1
                wbs_selected = []
                wbs_check_cols = st.columns(min(4, len(wbs_for_checklist)))
                for idx, wbs in enumerate(wbs_for_checklist):
                    with wbs_check_cols[idx % len(wbs_check_cols)]:
                        if st.checkbox(f"{wbs}", key=f"{sheet}_wbs_{wbs}", value=True):
                            wbs_selected.append(wbs)
                if not wbs_selected:
                    st.info("Please select at least one WBS component to proceed.")
                    continue
                # Map selected WBS to original WBS for analysis
                wbs_analysis_map = {}
                for wbs in wbs_selected:
                    if wbs in merge_groups:
                        for orig in merge_groups[wbs]:
                            wbs_analysis_map[orig] = wbs
                    else:
                        wbs_analysis_map[wbs] = wbs
                # Filter dataframe for selected WBS (including merged)
                df_selected = df[df[wbs_col].isin(wbs_analysis_map.keys())].copy()
                df_selected['WBS_Mapped'] = df_selected[wbs_col].map(wbs_analysis_map)
                # Mass entry section
                if len(wbs_selected) == 1:
                    st.markdown(f"#### {section_num}. Enter Mass for {wbs_selected[0]}")
                    section_num += 1
                    # Only one WBS, ask for a single mass input (total = individual)
                    wbs = wbs_selected[0]
                    unit = st.radio("Select the unit of your subsystem mass:", ["kg", "lbs"], key=f"unit_{sheet}")
                    mass = st.number_input(f"Enter the mass for {wbs}", min_value=0.0, step=0.1, key=f"mass_{sheet}_{wbs}")
                    if unit == "kg":
                        mass_lbs = mass * 2.20462
                    else:
                        mass_lbs = mass
                    st.caption(f"Mass entered: {mass} {unit} ({mass_lbs:.2f} lbs)")
                    wbs_mass_dict = {wbs: mass}
                    total_mass = mass
                    total_mass_lbs = mass_lbs
                    sum_wbs_mass = mass
                    # Show success if mass is entered
                    if mass > 0:
                        st.success(f"Mass for {wbs}: {mass} {unit} ({mass_lbs:.2f} lbs)")
                else:
                    st.markdown(f"#### {section_num}. Enter Mass for Each WBS Component (Sum must equal total mass)")
                    section_num += 1
                    # Ask for total mass and unit
                    unit = st.radio("Select the unit of your subsystem total mass:", ["kg", "lbs"], key=f"unit_{sheet}")
                    total_mass = st.number_input(f"Enter the total mass for {sheet}", min_value=0.0, step=0.1, key=f"total_mass_{sheet}")
                    if unit == "kg":
                        total_mass_lbs = total_mass * 2.20462
                    else:
                        total_mass_lbs = total_mass
                    st.caption(f"Total mass entered: {total_mass} {unit} ({total_mass_lbs:.2f} lbs)")
                    # Per-WBS mass entry
                    wbs_mass_dict = {}
                    wbs_mass_cols = st.columns(min(4, len(wbs_selected)))
                    for idx, wbs in enumerate(wbs_selected):
                        with wbs_mass_cols[idx % len(wbs_mass_cols)]:
                            wbs_mass_dict[wbs] = st.number_input(f"Mass for {wbs}", min_value=0.0, step=0.1, key=f"mass_{sheet}_{wbs}")
                    sum_wbs_mass = sum(wbs_mass_dict.values())
                    if abs(sum_wbs_mass - total_mass) > 1e-3:
                        st.warning(f"Sum of WBS masses ({sum_wbs_mass:.2f} {unit}) does not equal total mass ({total_mass:.2f} {unit})!")
                    else:
                        st.success(f"Sum of WBS masses matches total mass: {sum_wbs_mass:.2f} {unit}")
                # --- Inflation configuration ---
                st.markdown(f"#### {section_num}. Inflation Adjustment (Optional)")
                # Load inflation table
                infl_df = pd.read_excel('Inflation Table.xlsx', header=None)
                # Get year row and index row
                year_row = infl_df.iloc[5].tolist()[1:]
                index_row = infl_df.iloc[7].tolist()[1:]
                # Remove non-numeric years (e.g., 'TQ')
                year_index_pairs = [(y, idx) for y, idx in zip(year_row, index_row) if isinstance(y, (int, float))]
                years = [int(y) for y, _ in year_index_pairs]
                indices = [float(idx) for _, idx in year_index_pairs]
                year_to_index = dict(zip(years, indices))
                # Ask user for base year (template cost year) and target year
                st.info("You can adjust all costs for inflation using the NASA New Start Inflation Index.")
                # Set default index for base year to 1999 if present, else fallback to 2024 or 0
                if 1999 in years:
                    base_year_index = years.index(1999)
                elif 2024 in years:
                    base_year_index = years.index(2024)
                else:
                    base_year_index = 0
                base_year = st.selectbox("Which year are the costs in your template entered for?", years, index=base_year_index, key=f"base_year_{sheet}")
                target_year = st.selectbox("Which year do you want to escalate costs to?", years, index=years.index(2025) if 2025 in years else len(years)-1, key=f"target_year_{sheet}")
                # Compute inflation factor
                base_index = year_to_index.get(base_year, 1)
                target_index = year_to_index.get(target_year, 1)
                inflation_factor = target_index / base_index if base_index else 1
                st.caption(f"Inflation factor from {base_year} to {target_year}: {inflation_factor:.3f}")
                # Compute averages and price per pound for each WBS (merged or not)
                result_rows = []
                for wbs in wbs_selected:
                    # For merged groups, aggregate all original WBS in the group
                    wbs_rows = df_selected[df_selected['WBS_Mapped'] == wbs]
                    def safe_mean(col):
                        vals = pd.to_numeric(wbs_rows.get(col, pd.Series(dtype=float)), errors='coerce')
                        vals = vals.dropna()
                        return vals.mean() if not vals.empty else None
                    avg_weight = safe_mean("Higher Weight Range (lbs)")
                    avg_dd_cost = safe_mean("Higher D&D Cost Range")
                    avg_total_cost = safe_mean("Higher Total Cost Range")
                    avg_flight_unit_cost = safe_mean("Higher Flight Unit Cost Range")
                    count = len(wbs_rows)
                    # Use merged group name for mass entry
                    if len(wbs_selected) == 1:
                        # Only one group, use total/individual mass
                        if unit == "kg":
                            mass_lbs = total_mass * 2.20462
                        else:
                            mass_lbs = total_mass
                    else:
                        mass_lbs = wbs_mass_dict[wbs] * 2.20462 if unit == "kg" else wbs_mass_dict[wbs]
                    price_per_lb = avg_total_cost / avg_weight if avg_weight and avg_total_cost else None
                    est_price = price_per_lb * mass_lbs if price_per_lb else None
                    # Inflation-adjusted costs
                    adj_avg_dd_cost = avg_dd_cost * inflation_factor if avg_dd_cost else None
                    adj_avg_total_cost = avg_total_cost * inflation_factor if avg_total_cost else None
                    adj_est_price = est_price * inflation_factor if est_price else None
                    # New: Flight Unit Cost (reference and inflation-adjusted)
                    flight_unit_cost_new = ((avg_flight_unit_cost / avg_weight) * mass_lbs) if avg_flight_unit_cost and avg_weight else None
                    dd_cost_new = (avg_dd_cost / avg_weight * mass_lbs) if avg_dd_cost and avg_weight else None
                    adj_flight_unit_cost_new = flight_unit_cost_new * inflation_factor if flight_unit_cost_new else None
                    # Change Avg Higher Total Cost to be the sum of avg_dd_cost and avg_flight_unit_cost
                    result_rows.append({
                        "WBS": wbs,
                        "Count": count,
                        "Avg Higher Weight Range (lbs)": avg_weight,
                        "Avg Higher D&D Cost Range": avg_dd_cost,
                        "Avg Higher Flight Unit Cost": avg_flight_unit_cost,
                        "Avg Higher Total Cost": (avg_dd_cost if avg_dd_cost else 0) + (avg_flight_unit_cost if avg_flight_unit_cost else 0),
                        "User Mass (lbs)": mass_lbs,
                        "Est. Price (from hist.)": est_price,
                        "Total Cost": (flight_unit_cost_new if flight_unit_cost_new else 0) + (dd_cost_new if dd_cost_new else 0),  # Sum of flight unit cost and D&D cost for user system
                        "Flight Unit Cost per lbs": (avg_flight_unit_cost / avg_weight) if avg_flight_unit_cost and avg_weight else None,
                        "D&D Cost per lbs": (avg_dd_cost / avg_weight) if avg_dd_cost and avg_weight else None,
                        "D&D Cost": dd_cost_new,
                        "Flight Unit Cost (new, ref yr)": flight_unit_cost_new,
                        # "Price/lb (from hist.)": price_per_lb,  # Commented out as requested
                        # "Flight Unit Cost (new, ref yr)": (avg_flight_unit_cost / avg_weight * mass_lbs) if avg_flight_unit_cost and avg_weight else None,  # User system flight unit cost
                        f"Flight Unit Cost (new, {target_year})": adj_flight_unit_cost_new,
                        f"Adj. D&D Cost ({target_year})": adj_avg_dd_cost,
                        f"Adj. Total Cost ({target_year})": adj_avg_total_cost,
                        f"Adj. Est. Price ({target_year})": adj_est_price
                    })
                result_df = pd.DataFrame(result_rows)
                # Store result_df in session state for aggregation in Total Cost Breakdown
                if 'subsystem_results' not in st.session_state:
                    st.session_state['subsystem_results'] = {}
                st.session_state['subsystem_results'][sheet] = result_df
                # Rename columns for display: remove 'Range' everywhere
                result_df_display = result_df.rename(columns=lambda x: x.replace('Range', '').replace('range', '').replace('  ', ' ').replace('  ', ' ').strip())
                # Remove duplicate 'Total Cost Range' column, keep only 'Avg Higher Total Cost'
                hist_cols = [
                    "WBS", "Count", "Avg Higher Weight (lbs)", "Avg Higher D&D Cost", "Avg Higher Flight Unit Cost", "Avg Higher Total Cost",
                    f"Adj. D&D Cost ({target_year})", f"Adj. Total Cost ({target_year})"
                ]
                user_cols = [
                    "WBS", "User Mass (lbs)", "Total Cost", "D&D Cost", "Flight Unit Cost (new, ref yr)", "Flight Unit Cost per lbs", "D&D Cost per lbs"
                ]
                # Calculate inflation-adjusted per-lb costs for Table 3.3
                adj_flight_unit_cost_per_lbs = (result_df_display["Flight Unit Cost per lbs"] * inflation_factor) if "Flight Unit Cost per lbs" in result_df_display else None
                adj_dd_cost_per_lbs = (result_df_display["D&D Cost per lbs"] * inflation_factor) if "D&D Cost per lbs" in result_df_display else None
                # Add these columns to result_df_display for use in the inflation table
                if adj_flight_unit_cost_per_lbs is not None:
                    result_df_display[f"Adj. Flight Unit Cost per lbs ({target_year})"] = adj_flight_unit_cost_per_lbs
                if adj_dd_cost_per_lbs is not None:
                    result_df_display[f"Adj. D&D Cost per lbs ({target_year})"] = adj_dd_cost_per_lbs
                infl_cols = [
                    "WBS", f"Adj. Est. Price ({target_year})", f"Flight Unit Cost (new, {target_year})",
                    f"Adj. Flight Unit Cost per lbs ({target_year})", f"Adj. D&D Cost per lbs ({target_year})"
                ]
                st.markdown(f"#### {section_num}.1 Historical Data (Averages)")
                st.dataframe(result_df_display[hist_cols].set_index("WBS"), use_container_width=True)
                st.markdown(f"#### {section_num}.2 User Mass & Estimates")
                st.dataframe(result_df_display[user_cols].set_index("WBS"), use_container_width=True)
                st.markdown(f"#### {section_num}.3 Inflation Adjusted Estimates")
                st.dataframe(result_df_display[infl_cols].set_index("WBS"), use_container_width=True)
                # --- 3.3 Inflation Adjusted Estimates (EUR) ---
                EUR_CONV = 0.86  # Example: 1 USD = 0.86 EUR (update as needed)
                infl_cols_eur = [
                    col for col in infl_cols if col != "WBS"
                ]
                eur_df = result_df_display[["WBS"] + infl_cols_eur].copy()
                for col in infl_cols_eur:
                    eur_df[col] = eur_df[col].apply(lambda x: x * EUR_CONV if pd.notnull(x) else x)
                eur_df = eur_df.rename(columns={c: c + " (EUR)" for c in infl_cols_eur})
                st.markdown(f"#### {section_num}.4 Inflation Adjusted Estimates (EUR)")
                st.dataframe(eur_df.set_index("WBS"), use_container_width=True)
                # Store for Excel export
                st.session_state[f"eur_df_{sheet}"] = eur_df
                section_num += 1
                st.markdown("<div style='margin-top:1em;'></div>", unsafe_allow_html=True)
                selected_breakdown = st.selectbox(
                    "Select a WBS to view breakdown:",
                    options=["None"] + list(result_df["WBS"]),
                    key=f"breakdown_select_{sheet}"
                )
                if selected_breakdown and selected_breakdown != "None":
                    st.markdown(f"<div style='margin-top:1em; padding:1em; border-radius:8px; background:#f3f6fa; border:1px solid #e0e0e0; font-weight:bold; color:#222;'>Breakdown for WBS: <span style='color:#005fa3'>{selected_breakdown}</span></div>", unsafe_allow_html=True)
                    wbs_rows = df_selected[df_selected['WBS_Mapped'] == selected_breakdown]
                    st.dataframe(wbs_rows, use_container_width=True)
                with st.expander("Show/hide full data table", expanded=False):
                    st.dataframe(df)
                st.markdown("#### 4. Mass and Cost Trends Visualization")
                # Prepare data for plotting: show historical mass vs. cost for all selected/merged WBS
                import plotly.express as px
                # Combine all rows for selected/merged WBS
                plot_df = df_selected.copy()
                plot_df['User Mass (lbs)'] = plot_df['WBS_Mapped'].map({w: result_df.set_index('WBS').loc[w, 'User Mass (lbs)'] for w in result_df['WBS']})
                # Sliders for cost range
                min_cost = pd.to_numeric(plot_df['Higher Total Cost Range'], errors='coerce').min()
                max_cost = pd.to_numeric(plot_df['Higher Total Cost Range'], errors='coerce').max()
                if min_cost == max_cost or pd.isna(min_cost) or pd.isna(max_cost):
                    st.info(f"Only one unique value for Higher Total Cost (historical): {min_cost}")
                    cost_range = (min_cost, max_cost)
                else:
                    cost_range = st.slider(
                        "Select range for Higher Total Cost (historical)",
                        float(min_cost), float(max_cost), (float(min_cost), float(max_cost)), step=1.0,
                        key=f"cost_slider_{sheet}"
                    )
                filtered_plot_df = plot_df[(pd.to_numeric(plot_df['Higher Total Cost Range'], errors='coerce') >= cost_range[0]) & (pd.to_numeric(plot_df['Higher Total Cost Range'], errors='coerce') <= cost_range[1])]
                # Plot mass vs. cost for each WBS_Mapped
                fig = px.scatter(filtered_plot_df, x='Higher Weight Range (lbs)', y='Higher Total Cost Range', color='WBS_Mapped',
                    hover_data=['Mission', 'WBS_Mapped', 'Higher D&D Cost Range', 'Higher Total Cost Range', 'Higher Weight Range (lbs)'],
                    title='Historical Mass vs. Total Cost by WBS', labels={'Higher Weight Range (lbs)': 'Mass (lbs)', 'Higher Total Cost Range': 'Total Cost'})
                st.plotly_chart(fig, use_container_width=True)
                # Optionally, show cost per unit mass trend
                filtered_plot_df['Cost per lb'] = pd.to_numeric(filtered_plot_df['Higher Total Cost Range'], errors='coerce') / pd.to_numeric(filtered_plot_df['Higher Weight Range (lbs)'], errors='coerce')
                fig2 = px.scatter(filtered_plot_df, x='Higher Weight Range (lbs)', y='Cost per lb', color='WBS_Mapped',
                    hover_data=['Mission', 'WBS_Mapped', 'Higher D&D Cost Range', 'Higher Total Cost Range', 'Higher Weight Range (lbs)'],
                    title='Historical Mass vs. Cost per Unit Mass by WBS', labels={'Higher Weight Range (lbs)': 'Mass (lbs)', 'Cost per lb': 'Cost per lb'})
                st.plotly_chart(fig2, use_container_width=True)
                st.markdown("#### 4. Mass vs. Cost Line Plot (Interactive)")
                # Prepare data for line plot: x=mass, y=costs (3 lines)
                import plotly.graph_objects as go
                # Use all selected/merged WBS rows
                plot_df = df_selected.copy()
                plot_df['Higher Weight Range (lbs)'] = pd.to_numeric(plot_df['Higher Weight Range (lbs)'], errors='coerce')
                plot_df['Higher D&D Cost Range'] = pd.to_numeric(plot_df['Higher D&D Cost Range'], errors='coerce')
                plot_df['Higher Flight Unit Cost Range'] = pd.to_numeric(plot_df['Higher Flight Unit Cost Range'], errors='coerce')
                plot_df['Higher Total Cost Range'] = pd.to_numeric(plot_df['Higher Total Cost Range'], errors='coerce')
                # Remove rows with missing mass
                plot_df = plot_df.dropna(subset=['Higher Weight Range (lbs)'])
                # Slider for mass range
                min_mass = float(plot_df['Higher Weight Range (lbs)'].min() or 0)
                max_mass = float(plot_df['Higher Weight Range (lbs)'].max() or 1)
                if min_mass == max_mass or pd.isna(min_mass) or pd.isna(max_mass):
                    st.info(f"Only one unique value for Mass (lbs): {min_mass}")
                    mass_range = (min_mass, max_mass)
                else:
                    mass_range = st.slider(
                        "Select mass (lbs) range to display",
                        min_mass, max_mass, (min_mass, max_mass), step=1.0,
                        key=f"mass_slider_{sheet}"
                    )
                filtered_plot_df = plot_df[(plot_df['Higher Weight Range (lbs)'] >= mass_range[0]) & (plot_df['Higher Weight Range (lbs)'] <= mass_range[1])]
                filtered_plot_df = filtered_plot_df.sort_values('Higher Weight Range (lbs)')
                fig = go.Figure()
                fig.add_trace(go.Scatter(x=filtered_plot_df['Higher Weight Range (lbs)'], y=filtered_plot_df['Higher D&D Cost Range'],
                                         mode='lines+markers', name='Higher D&D Cost'))
                fig.add_trace(go.Scatter(x=filtered_plot_df['Higher Weight Range (lbs)'], y=filtered_plot_df['Higher Flight Unit Cost Range'],
                                         mode='lines+markers', name='Higher Flight Unit Cost'))
                fig.add_trace(go.Scatter(x=filtered_plot_df['Higher Weight Range (lbs)'], y=filtered_plot_df['Higher Total Cost Range'],
                                         mode='lines+markers', name='Higher Total Cost'))
                fig.update_layout(title='Mass vs. Cost (Historical Data)',
                                  xaxis_title='Mass (lbs)',
                                  yaxis_title='Cost',
                                  legend_title='Cost Type',
                                  hovermode='x unified')
                st.plotly_chart(fig, use_container_width=True)
        # --- Place Export to Excel button at the bottom, always visible ---
        st.markdown("---")
        st.subheader("Export to Excel")
        st.caption("Download an Excel file containing all your cost analysis results and breakdowns. Use this after you have completed your analysis.")
        export_clicked = st.button("Export to Excel", key="export_to_excel")
        if export_clicked:
            uploaded = st.session_state.get('uploaded_file')
            if not uploaded:
                st.warning("Please upload a filled template before exporting to Excel.")
            else:
                import io
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                output = io.BytesIO()
                wb = Workbook()
                wb.remove(wb.active)
                subsystem_results = st.session_state.get('subsystem_results', {})
                # Define styles
                header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                header_align = Alignment(horizontal="center", vertical="center")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                alt_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                # Color palette for sheets
                sheet_colors = ["4F81BD", "C0504D", "9BBB59", "8064A2", "F79646", "2C4D75", "1F497D", "E46C0A", "00B050", "7030A0"]
                for idx, (sheet, result_df) in enumerate(subsystem_results.items()):
                    ws = wb.create_sheet(title=sheet[:31])
                    color = sheet_colors[idx % len(sheet_colors)]
                    sheet_header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    # Write headers
                    for c_idx, col in enumerate(result_df.columns, 1):
                        cell = ws.cell(row=1, column=c_idx, value=col)
                        cell.fill = sheet_header_fill
                        cell.font = header_font
                        cell.alignment = header_align
                        cell.border = thin_border
                    # Write data rows with alternating fill
                    for r_idx, row in enumerate(result_df.reset_index(drop=True).itertuples(index=False), 2):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=value)
                            cell.border = thin_border
                            if r_idx % 2 == 0:
                                cell.fill = alt_fill
                    # Auto-fit columns
                    for col in ws.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
                    # --- Add EUR table below main table ---
                    eur_df = st.session_state.get(f"eur_df_{sheet}")
                    if eur_df is not None:
                        start_row = ws.max_row + 2
                        for c_idx, col in enumerate(eur_df.columns, 1):
                            cell = ws.cell(row=start_row, column=c_idx, value=col)
                            cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                            cell.font = header_font
                            cell.alignment = header_align
                            cell.border = thin_border
                        for r_idx, row in enumerate(eur_df.reset_index(drop=True).itertuples(index=False), start_row+1):
                            for c_idx, value in enumerate(row, 1):
                                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                                cell.border = thin_border
                                if r_idx % 2 == 0:
                                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    # Auto-fit columns again
                    for col in ws.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
                # Total Cost Breakdown sheet (distinct style)
                if 'subsystem_results' in st.session_state:
                    ws = wb.create_sheet(title="Total Cost Breakdown")
                    user_mass_df = st.session_state.get('user_mass_df')
                    infl_df = st.session_state.get('infl_df')
                    breakdown_header_fill = PatternFill(start_color="005fa3", end_color="005fa3", fill_type="solid")
                    breakdown_alt_fill = PatternFill(start_color="E3F0FF", end_color="E3F0FF", fill_type="solid")
                    if user_mass_df is not None:
                        for c_idx, col in enumerate(user_mass_df.columns, 1):
                            cell = ws.cell(row=1, column=c_idx, value=col)
                            cell.fill = breakdown_header_fill
                            cell.font = header_font
                            cell.alignment = header_align
                            cell.border = thin_border
                        for r_idx, row in enumerate(user_mass_df.reset_index(drop=True).itertuples(index=False), 2):
                            for c_idx, value in enumerate(row, 1):
                                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                                cell.border = thin_border
                                if r_idx % 2 == 0:
                                    cell.fill = breakdown_alt_fill
                    if infl_df is not None:
                        start_row = ws.max_row + 2
                        for c_idx, col in enumerate(infl_df.columns, 1):
                            cell = ws.cell(row=start_row, column=c_idx, value=col)
                            cell.fill = breakdown_header_fill
                            cell.font = header_font
                            cell.alignment = header_align
                            cell.border = thin_border
                        for r_idx, row in enumerate(infl_df.reset_index(drop=True).itertuples(index=False), start_row+1):
                            for c_idx, value in enumerate(row, 1):
                                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                                cell.border = thin_border
                                if r_idx % 2 == 0:
                                    cell.fill = breakdown_alt_fill
                    # Auto-fit columns
                    for col in ws.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
                wb.save(output)
                st.download_button(
                    label="Download Cost Analysis Excel",
                    data=output.getvalue(),
                    file_name="CostSpirits_Cost_Analysis.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
if __name__ == "__main__":
    main()
